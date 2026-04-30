#!/usr/bin/env python3
"""Replay cache benchmark for ZenMux and OpenRouter.

Each platform keeps its own multi-turn history.
Every item in QUESTIONS is used as the next user turn.
The platform's real assistant reply is appended back into that platform's
history before the next round.
"""

import os
import copy
import json
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

OUTPUT_DIR = "cache-benchmark-replay-" + datetime.now().strftime("%Y%m%d_%H%M%S")
REQUEST_TIMEOUT_SECONDS = 60 * 20

QUESTIONS = [
                [
                    "What problem does dynamic programming (DP) actually solve",
                    "What is the relationship between DP and recursion, and why is DP more efficient",
                    "Why does DP need a memoization table, isn't plain recursion enough",
                    "Explain in detail what a DP state transition equation is",
                    "What does optimal substructure mean in DP",
                    "What does overlapping subproblems mean in DP",
                    "Explain optimal substructure and overlapping subproblems in DP to a beginner",
                    "What is the difference in state transitions between 0/1 knapsack and unbounded knapsack",
                    "What is the optimal substructure of the stair climbing problem",
                    "How do you define the DP state for the Longest Increasing Subsequence problem",
                    "How do you distinguish between dynamic programming and greedy algorithms, and what scenarios suit each",
                    "In a real interview, how do you tell whether a problem can be solved with DP",
                    "What techniques reduce space complexity from O(n²) to O(n)",
                ],
                [
                    "For a first-time renovation, should I choose half-package or full-package service",
                    "What is the difference between half-package and full-package renovation",
                    "Which is more suitable for a first-timer, half-package or full-package",
                    "Explain in detail the difference between hard decoration and soft decoration",
                    "How should the budget be allocated between hard decoration and soft decoration",
                    "Which can be changed later, hard decoration or soft decoration",
                    "Where are the most common scams in electrical and plumbing renovation",
                    "Why is there such a big price difference between half-package and full-package for electrical and plumbing work",
                    "What is the difference between light luxury style and creamy style in hard decoration",
                    "Explain the difference between light luxury style and creamy style to parents",
                    "Should I choose solid wood, engineered wood, or laminate flooring",
                    "How do I choose eco-friendly panels for whole-house custom furniture",
                    "Why do renovation projects always go over budget, and how can I control costs",
                    "How to scientifically deal with formaldehyde anxiety during renovation",
                ],
                [
                    "In 2008, if you had 1 million, should you have put it in the bank or bought property",
                    "During the 2008 financial crisis, if you had 1 million, should you have put it in the bank or bought gold",
                    "Explain in detail the causes of the 2008 financial crisis",
                    "Why did the bankruptcy of Lehman Brothers affect the entire world",
                    "What did liquidity crisis mean during the 2008 financial crisis",
                    "Explain the liquidity crisis of the 2008 financial crisis to a friend who doesn't understand finance",
                    "Could buying gold in 2008 preserve value",
                    "Which was worse at the time, buying A-shares or buying US stocks",
                    "Was buying real estate at the bottom in 2008 an opportunity or a trap",
                    "What exactly was the Four Trillion stimulus plan",
                    "What impact did the Four Trillion stimulus plan have on housing prices during the 2008 financial crisis",
                    "Why was cash king during the 2008 financial crisis",
                    "Were fixed deposits and government bonds safe during the crisis",
                    "If someone had bought Bitcoin at that time, how much would it be worth now (historical hypothetical)",
                    "Looking back at the 2008 financial crisis, what would have been the best move with 1 million",
                ],
            ]

MODEL = "openai/gpt-5.4"
PROVIDER = "openai"

PLATFORMS = [
    {
        "name": "zenmux",
        "base_url": "https://zenmux.ai/api/v1",
        "header_request_id": "x-zenmux-requestid",
        # TODO
        "api_key": "",
        "model": MODEL,
        "ext_body": {
            "provider": {"only": [PROVIDER], "allow_fallbacks": False},
        },
    },
    {
        "name": "openrouter",
        "base_url": "https://openrouter.ai/api/v1",
        "header_request_id": "x-generation-id",
        # TODO
        "api_key": "",
        "model": MODEL,
        "ext_body": {
            "provider": {"only": [PROVIDER], "allow_fallbacks": False},
        },
    },
]


def extract_assistant_text(body):
    content = (body.get("choices") or [{}])[0].get("message", {}).get("content", "")
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts = []
        for item in content:
            if isinstance(item, dict) and item.get("type") == "text":
                parts.append(str(item.get("text", "")))
        return "".join(parts)
    return ""


def extract_provider(platform):
    provider = ((platform.get("ext_body") or {}).get("provider") or {}).get("only") or []
    if provider:
        return str(provider[0])
    return ""


def send_chat_completion(platform, messages, round_index):
    payload = platform["ext_body"] | {
        "model": platform["model"],
        "messages": messages,
        "stream": False,
        "reasoning": {"effort": "high"},
    }
    headers = {
        "Authorization": f"Bearer {platform['api_key']}",
        "Content-Type": "application/json",
    }

    status_code = 0
    error = ""
    body = {}
    request_id = ""
    start = time.perf_counter()

    try:
        response = requests.post(
            f"{platform['base_url'].rstrip('/')}/chat/completions",
            headers=headers,
            json=payload,
            timeout=REQUEST_TIMEOUT_SECONDS,
        )
        status_code = response.status_code
        request_id = response.headers.get(platform.get('header_request_id'))
        try:
            body = response.json() or {}
        except ValueError:
            error = "response is not valid json"
    except requests.RequestException as exc:
        print(f"[round {round_index}] platform:{platform['name']} has error")
        error = str(exc)

    usage = body.get("usage") or {}
    details = usage.get("prompt_tokens_details") or {}

    prompt_tokens = usage.get("prompt_tokens") or 0
    completion_tokens = usage.get("completion_tokens") or 0
    total_tokens = usage.get("total_tokens") or 0
    cached_tokens = details.get("cached_tokens")

    return {
        "mode": "replay",
        "platform": platform["name"],
        "model": platform["model"],
        "request_id": request_id,
        "status_code": status_code,
        "error": error,
        "latency_ms": round((time.perf_counter() - start) * 1000, 2),
        "prompt_tokens": prompt_tokens,
        "completion_tokens": completion_tokens,
        "total_tokens": total_tokens,
        "cached_tokens": cached_tokens,
        "token_hit_rate": (cached_tokens / prompt_tokens) if prompt_tokens else 0.0,
        "assistant_text": extract_assistant_text(body),
    }


def build_summary_rows(rows):
    summary_rows = []
    for row in rows:
        usage_token = {
            "prompt_tokens": row.get("prompt_tokens", 0),
            "completion_tokens": row.get("completion_tokens", 0),
            "total_tokens": row.get("total_tokens", 0),
            "cached_tokens": row.get("cached_tokens", 0),
        }
        summary_rows.append(
            {
                "model": row.get("model", ""),
                "provider": row.get("provider", ""),
                "question_group": row.get("question_group", ""),
                "round": row.get("round", ""),
                "user_question": row.get("user_question", ""),
                "platform": row.get("platform", ""),
                "request_id": row.get("request_id", ""),
                "usage_token_json": json.dumps(usage_token, ensure_ascii=False),
                "cache_hit_rate": row.get("token_hit_rate", 0.0),
            }
        )

    return sorted(
        summary_rows,
        key=lambda summary_row: (
            summary_row["model"],
            summary_row["provider"],
            summary_row["question_group"],
            summary_row["round"],
            summary_row["user_question"],
            summary_row["platform"],
        ),
    )


def merge_by_fields(sheet, rows, key_fields, merge_columns):
    if not rows:
        return

    start_idx = 0
    while start_idx < len(rows):
        end_idx = start_idx
        current_key = tuple(rows[start_idx][field] for field in key_fields)
        while end_idx + 1 < len(rows):
            next_key = tuple(rows[end_idx + 1][field] for field in key_fields)
            if next_key != current_key:
                break
            end_idx += 1

        if end_idx > start_idx:
            start_row = start_idx + 2
            end_row = end_idx + 2
            for col in merge_columns:
                sheet.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row,
                    end_column=col,
                )
                sheet.cell(start_row, col).alignment = Alignment(
                    wrap_text=True,
                    vertical="center",
                )

        start_idx = end_idx + 1


def build_cache_summary_workbook(summary_rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "cache_summary"

    headers = [
        "model",
        "provider",
        "question_group",
        "round",
        "user_question",
        "platform",
        "request_id",
        "usage_token_json",
        "cache_hit_rate",
    ]

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    for row in summary_rows:
        sheet.append([row[header] for header in headers])

    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    merge_by_fields(sheet, summary_rows, ["model"], [1])
    merge_by_fields(sheet, summary_rows, ["model", "provider"], [2])
    merge_by_fields(sheet, summary_rows, ["model", "provider", "question_group"], [3])
    merge_by_fields(
        sheet,
        summary_rows,
        ["model", "provider", "question_group", "round", "user_question"],
        [4, 5],
    )

    for cell in sheet["I"][1:]:
        cell.number_format = "0.00%"

    widths = {
        "A": 24,
        "B": 14,
        "C": 14,
        "D": 10,
        "E": 50,
        "F": 14,
        "G": 24,
        "H": 55,
        "I": 14,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    sheet.freeze_panes = "A2"
    return workbook


def run_question_group(question_group, question_group_index):
    rows = []
    histories = {platform["name"]: [] for platform in PLATFORMS}

    for round_index, question in enumerate(question_group, start=1):
        print(f"[round {round_index}] {question}")
        for platform in PLATFORMS:
            print(f"[round {round_index}] platform:{platform['name']}")
            histories[platform["name"]].append({"role": "user", "content": question})
            messages = copy.deepcopy(histories[platform["name"]])
            row = send_chat_completion(platform, messages, round_index)
            assistant_text = row.pop("assistant_text", "")
            row["round"] = round_index
            row["user_question"] = question
            row["question_group"] = question_group_index
            row["provider"] = extract_provider(platform)
            rows.append(row)
            histories[platform["name"]].append(
                {"role": "assistant", "content": assistant_text}
            )

    return rows


def save_report(output_dir, rows):
    workbook = build_cache_summary_workbook(build_summary_rows(rows))
    workbook.save(output_dir / f"cache_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


def main():
    output_dir = Path(OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_rows = []
    question_group_index = 0
    for question_group in QUESTIONS:
        if not question_group:
            continue
        question_group_index += 1
        rows = run_question_group(question_group, question_group_index)
        all_rows.extend(rows)

        resp_filepath = os.path.join(OUTPUT_DIR, f"{question_group_index}-resp.json")
        with open(resp_filepath, "w", encoding="utf-8") as f:
            f.write(json.dumps(rows, ensure_ascii=False, indent=2))

    save_report(output_dir, all_rows)


if __name__ == "__main__":
    main()
